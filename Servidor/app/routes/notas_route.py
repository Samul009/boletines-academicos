# routers/notas_route.py
from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List
from datetime import datetime, date
from ..core.database import get_db
from ..core.permissions import require_permission
from ..models.models import (
    DocenteAsignatura, Grupo, Asignatura, AnioLectivo, PeriodoAcademico,
    Matricula, Persona, Calificacion, Falla, Grado, Usuario
)
from ..models.notas_schemas import (
    DocenteClaseSchema, EstudianteNotaSchema, DashboardDocenteSchema
)
from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import Font

router = APIRouter(prefix="/notas", tags=["Notas, Boletines y Asistencia"])


# =======================
# 1. DASHBOARD DOCENTE
# =======================
@router.get("/dashboard", response_model=DashboardDocenteSchema)
def dashboard_docente(
    id_usuario: int = Query(..., description="ID del docente"),
    db: Session = Depends(get_db),
    user = Depends(require_permission("/nota", "ver"))
):
    asignaturas = db.query(DocenteAsignatura).filter(
        DocenteAsignatura.id_usuario_docente == id_usuario,
        DocenteAsignatura.fecha_eliminacion.is_(None)
    ).join(Grupo).join(Asignatura).join(AnioLectivo).join(Grado).all()

    clases = [
        DocenteClaseSchema(
            id_docente_asignatura=c.id_docente_asignatura,
            id_asignatura=c.id_asignatura,
            nombre_asignatura=c.asignatura.nombre_asignatura,
            id_grupo=c.id_grupo,
            codigo_grupo=c.grupo.codigo_grupo,
            nombre_grado=c.grupo.grado.nombre_grado,
            id_anio_lectivo=c.id_anio_lectivo,
            anio=c.anio_lectivo.anio
        ) for c in asignaturas
    ]

    grupos_dirigidos = db.query(Grupo).filter(
        Grupo.id_usuario_director == id_usuario,
        Grupo.fecha_eliminacion.is_(None)
    ).join(Grado).join(AnioLectivo).all()

    director = [
        {
            "id_grupo": g.id_grupo,
            "codigo_grupo": g.codigo_grupo,
            "nombre_grado": g.grado.nombre_grado,
            "anio": g.anio_lectivo.anio
        } for g in grupos_dirigidos
    ]

    return DashboardDocenteSchema(asignaturas=clases, director_de_grupo=director)


# =======================
# 2. ESTUDIANTES + NOTAS + FALLAS
# =======================
@router.get("/clase/{id_docente_asignatura}/periodo/{id_periodo}", response_model=List[EstudianteNotaSchema])
def estudiantes_con_nota_periodo(
    id_docente_asignatura: int,
    id_periodo: int,
    db: Session = Depends(get_db),
    user = Depends(require_permission("/nota", "ver"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da or da.fecha_eliminacion: raise HTTPException(404, "Asignación no encontrada")

    periodo = db.get(PeriodoAcademico, id_periodo)
    if not periodo or periodo.id_anio_lectivo != da.id_anio_lectivo:
        raise HTTPException(400, "Período inválido")

    query = text("""
        SELECT 
            p.id_persona, p.nombre, p.apellido, p.foto,
            c.calificacion_numerica,
            COUNT(f.id_falla) as total_fallas,
            SUM(CASE WHEN f.es_justificada THEN 1 ELSE 0 END) as justificadas,
            SUM(CASE WHEN NOT f.es_justificada THEN 1 ELSE 0 END) as injustificadas
        FROM persona p
        JOIN matricula m ON m.id_persona = p.id_persona
        LEFT JOIN calificacion c ON c.id_persona = p.id_persona 
            AND c.id_asignatura = :asig AND c.id_periodo = :periodo 
            AND c.id_anio_lectivo = :anio AND c.fecha_eliminacion IS NULL
        LEFT JOIN falla f ON f.id_persona = p.id_persona 
            AND f.id_asignatura = :asig AND f.fecha_eliminacion IS NULL
        WHERE m.id_grupo = :grupo AND m.id_anio_lectivo = :anio AND m.activo = TRUE
          AND p.fecha_eliminacion IS NULL
        GROUP BY p.id_persona, p.nombre, p.apellido, p.foto, c.calificacion_numerica
    """)

    result = db.execute(query, {
        "asig": da.id_asignatura, "periodo": id_periodo,
        "anio": da.id_anio_lectivo, "grupo": da.id_grupo
    }).fetchall()

    return [
        EstudianteNotaSchema(
            id_persona=r.id_persona,
            nombre=r.nombre,
            apellido=r.apellido,
            foto=r.foto,
            nota_existente=float(r.calificacion_numerica) if r.calificacion_numerica else None,
            total_fallas=r.total_fallas,
            fallas_justificadas=r.justificadas,
            fallas_injustificadas=r.injustificadas
        ) for r in result
    ]


# =======================
# 3. EXPORTAR PLANTILLA NOTAS (CON CÉDULA)
# =======================
@router.get("/exportar-plantilla/{id_docente_asignatura}/periodo/{id_periodo}")
def exportar_plantilla_excel(
    id_docente_asignatura: int,
    id_periodo: int,
    db: Session = Depends(get_db),
    user = Depends(require_permission("/nota", "exportar"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da: raise HTTPException(404, "No encontrada")

    periodo = db.get(PeriodoAcademico, id_periodo)
    if not periodo or periodo.id_anio_lectivo != da.id_anio_lectivo:
        raise HTTPException(400, "Período inválido")

    # Construir query según si id_grupo es NULL o tiene valor
    if da.id_grupo is None:
        # Obtener todos los grupos del grado
        grupos_query = db.execute(text("""
            SELECT id_grupo FROM grupo 
            WHERE id_grado = :grado AND id_anio_lectivo = :anio AND fecha_eliminacion IS NULL
        """), {"grado": da.id_grado, "anio": da.id_anio_lectivo}).fetchall()
        grupos_ids = [g[0] for g in grupos_query]
        
        estudiantes = db.execute(text("""
            SELECT p.apellido, p.nombre, p.numero_identificacion, c.calificacion_numerica
            FROM persona p
            JOIN matricula m ON m.id_persona = p.id_persona
            LEFT JOIN calificacion c ON c.id_persona = p.id_persona 
                AND c.id_asignatura = :asig AND c.id_periodo = :periodo 
                AND c.id_anio_lectivo = :anio
            WHERE m.id_grupo IN :grupos AND m.id_anio_lectivo = :anio AND m.activo = TRUE
            ORDER BY p.apellido
        """), {
            "asig": da.id_asignatura, "periodo": id_periodo,
            "anio": da.id_anio_lectivo, "grupos": tuple(grupos_ids)
        }).fetchall()
    else:
        estudiantes = db.execute(text("""
            SELECT p.apellido, p.nombre, p.numero_identificacion, c.calificacion_numerica
            FROM persona p
            JOIN matricula m ON m.id_persona = p.id_persona
            LEFT JOIN calificacion c ON c.id_persona = p.id_persona 
                AND c.id_asignatura = :asig AND c.id_periodo = :periodo 
                AND c.id_anio_lectivo = :anio
            WHERE m.id_grupo = :grupo AND m.id_anio_lectivo = :anio AND m.activo = TRUE
            ORDER BY p.apellido
        """), {
            "asig": da.id_asignatura, "periodo": id_periodo,
            "anio": da.id_anio_lectivo, "grupo": da.id_grupo
        }).fetchall()

    data = []
    for i, est in enumerate(estudiantes):
        data.append({
            "N°": i + 1,
            "Cédula": est.numero_identificacion or "",
            "Apellido": est.apellido,
            "Nombre": est.nombre,
            "Nota": est.calificacion_numerica or ""
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Notas')
        ws = writer.sheets['Notas']
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

    output.seek(0)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Notas_{da.asignatura.nombre_asignatura}_{periodo.nombre_periodo}.xlsx"}
    )


# =======================
# 4. IMPORTAR NOTAS (CON CÉDULA)
# =======================
@router.post("/importar-notas")
async def importar_notas(
    file: UploadFile = File(...),
    id_docente_asignatura: int = Query(...),
    id_periodo: int = Query(...),
    db: Session = Depends(get_db),
    user = Depends(require_permission("/nota", "importar"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da: raise HTTPException(404, "No encontrada")

    contents = await file.read()
    df = pd.read_excel(BytesIO(contents))

    resultados = []
    for _, row in df.iterrows():
        cedula = str(row.get("Cédula", "")).strip()
        apellido = str(row.get("Apellido", "")).strip()
        nombre = str(row.get("Nombre", "")).strip()

        if not cedula:
            resultados.append({"error": f"Cédula faltante: {apellido} {nombre}"})
            continue

        persona = db.query(Persona).join(Matricula).filter(
            Persona.numero_identificacion == cedula,
            Matricula.id_grupo == da.id_grupo,
            Matricula.id_anio_lectivo == da.id_anio_lectivo,
            Matricula.activo == True
        ).first()

        if not persona:
            resultados.append({"error": f"No encontrado: {cedula}"})
            continue

        if pd.notna(row.get("Nota")):
            try:
                nota = float(row["Nota"])
                if not 0.0 <= nota <= 5.0:
                    resultados.append({"error": f"Nota inválida: {nota}"})
                    continue

                cal = db.query(Calificacion).filter(
                    Calificacion.id_persona == persona.id_persona,
                    Calificacion.id_asignatura == da.id_asignatura,
                    Calificacion.id_periodo == id_periodo,
                    Calificacion.id_anio_lectivo == da.id_anio_lectivo
                ).first()

                if cal:
                    cal.calificacion_numerica = nota
                    cal.fecha_actualizacion = datetime.now()
                else:
                    cal = Calificacion(
                        id_persona=persona.id_persona,
                        id_asignatura=da.id_asignatura,
                        id_periodo=id_periodo,
                        id_anio_lectivo=da.id_anio_lectivo,
                        id_usuario=da.id_usuario_docente,
                        calificacion_numerica=nota
                    )
                    db.add(cal)
            except:
                resultados.append({"error": f"Nota no numérica: {row.get('Nota')}"})
                continue

        resultados.append({"estudiante": f"{cedula} - {apellido} {nombre}", "status": "OK"})

    db.commit()
    return {"resultados": resultados}


# =======================
# 5. EXPORTAR ASISTENCIA (CORREGIDO: docente por ID)
# =======================
# CORREGIDO: exportar_asistencia
@router.get("/exportar-asistencia/{id_docente_asignatura}/mes/{mes}/anio/{anio}")
def exportar_asistencia(
    id_docente_asignatura: int,
    mes: int,
    anio: int,
    db: Session = Depends(get_db),
    user = Depends(require_permission("/asistencia", "exportar"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da: raise HTTPException(404, "No encontrada")
    if not 1 <= mes <= 12: raise HTTPException(400, "Mes inválido")

    docente = db.get(Usuario, da.id_usuario_docente)
    if not docente: raise HTTPException(404, "Docente no encontrado")

    estudiantes = db.execute(text("""
        SELECT p.apellido, p.nombre
        FROM persona p
        JOIN matricula m ON m.id_persona = p.id_persona
        WHERE m.id_grupo = :grupo AND m.id_anio_lectivo = :anio_lectivo AND m.activo = TRUE
        ORDER BY p.apellido, p.nombre
    """), {"grupo": da.id_grupo, "anio_lectivo": da.id_anio_lectivo}).fetchall()

    # === CREAR DATAFRAME SIN TÍTULOS ===
    data = []
    for i, est in enumerate(estudiantes):
        row = [i + 1, f"{est.apellido} {est.nombre}"] + [""] * 31
        data.append(row)

    df = pd.DataFrame(data, columns=["N°", "NOMBRE COMPLETO"] + [str(d) for d in range(1, 32)])
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir sin encabezado
        df.to_excel(writer, index=False, sheet_name='Asistencia', header=False, startrow=3)

        worksheet = writer.sheets['Asistencia']

        # === TÍTULOS MANUALES ===
        worksheet['A1'] = f"JORNADA: Mañana - Grado: {da.grupo.grado.nombre_grado} - Docente: {docente.persona.nombre} {docente.persona.apellido}"
        worksheet.merge_cells('A1:AH1')
        worksheet['A1'].font = Font(bold=True, size=12)

        worksheet['A2'] = f"MES: {mes:02d}/{anio}"
        worksheet.merge_cells('A2:AH2')
        worksheet['A2'].font = Font(bold=True)

        # === ENCABEZADOS DE COLUMNAS (día 1 a 31) ===
        for dia in range(1, 32):
            col_letter = openpyxl.utils.get_column_letter(dia + 2)  # A=1, B=2 → C=3
            worksheet[f"{col_letter}3"] = str(dia)
            worksheet[f"{col_letter}3"].font = Font(bold=True)

        # === ANCHOS ===
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 28
        for col in range(3, 34):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3.5

    output.seek(0)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Asistencia_{da.asignatura.nombre_asignatura}_{mes:02d}_{anio}.xlsx"}
    )

# =======================
# 6. IMPORTAR ASISTENCIA (100% FUNCIONAL)
# =======================
@router.post("/importar-asistencia")
async def importar_asistencia(
    file: UploadFile = File(...),
    id_docente_asignatura: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    user = Depends(require_permission("/asistencia", "importar"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da: raise HTTPException(404, "No encontrada")

    contents = await file.read()

    # === LEER SIN ENCABEZADOS + SALTAR FILAS 1 y 2 ===
    try:
        df = pd.read_excel(BytesIO(contents), header=None, skiprows=3)
        print("FILAS LEÍDAS:", len(df))
        print("PRIMERA FILA:", df.iloc[0].to_dict())
    except Exception as e:
        raise HTTPException(400, f"Error leyendo Excel: {str(e)}")

    if df.empty:
        raise HTTPException(400, "El archivo está vacío o mal formado")

    resultados = []
    fallas_agregadas = 0

    for _, row in df.iterrows():
        # === COLUMNA 0 = N° | COLUMNA 1 = NOMBRE COMPLETO ===
        try:
            numero = row.iloc[0]
            nombre_completo = str(row.iloc[1]).strip()
        except IndexError:
            continue

        if pd.isna(numero) or pd.isna(nombre_completo) or nombre_completo.lower() == "nan":
            continue

        print(f"PROCESANDO: {nombre_completo}")

        # === BUSCAR ESTUDIANTE ===
        partes = nombre_completo.split()
        if len(partes) < 2:
            resultados.append({"error": f"Nombre inválido: {nombre_completo}"})
            continue

        apellido = partes[0]
        nombre = " ".join(partes[1:])

        persona = db.query(Persona).join(Matricula).filter(
            Persona.apellido.ilike(apellido),
            Persona.nombre.ilike(f"%{nombre}%"),
            Matricula.id_grupo == da.id_grupo,
            Matricula.id_anio_lectivo == da.id_anio_lectivo,
            Matricula.activo == True
        ).first()

        if not persona:
            resultados.append({"error": f"No encontrado: {nombre_completo}"})
            continue

        # === PROCESAR DÍAS (columnas 2 en adelante) ===
        for dia_idx in range(2, len(row)):
            if dia_idx >= len(row) or pd.isna(row.iloc[dia_idx]):
                continue

            marca_raw = str(row.iloc[dia_idx]).strip().upper()
            
            # CORREGIR 'N' → 'F'
            marca = 'F' if marca_raw == 'N' else marca_raw

            if marca not in ['S', 'F', 'J']:
                continue

            dia = dia_idx - 1  # columna 2 = día 1
            try:
                fecha = date(anio, mes, dia)
            except ValueError:
                continue

            # GUARDAR SOLO SI ES F O J
            if marca in ['F', 'J']:
                existe = db.query(Falla).filter(
                    Falla.id_persona == persona.id_persona,
                    Falla.id_asignatura == da.id_asignatura,
                    Falla.fecha_falla == fecha
                ).first()

                if not existe:
                    db.add(Falla(
                        id_persona=persona.id_persona,
                        id_asignatura=da.id_asignatura,
                        fecha_falla=fecha,
                        es_justificada=(marca == 'J')
                    ))
                    fallas_agregadas += 1

        resultados.append({"estudiante": nombre_completo, "status": "OK"})

    db.commit()

    return {
        "resultados": resultados,
        "total_fallas_registradas": fallas_agregadas,
        "mensaje": f"Se procesaron {len(resultados)} estudiantes"
    }


# =======================
# 7. DATOS PARA BOLETÍN
# =======================
@router.get("/datos-boletin/{id_docente_asignatura}/periodo/{id_periodo}")
def datos_boletin(
    id_docente_asignatura: int,
    id_periodo: int,
    db: Session = Depends(get_db),
    user = Depends(require_permission("/boletin", "ver"))
):
    da = db.get(DocenteAsignatura, id_docente_asignatura)
    if not da: raise HTTPException(404, "No encontrada")

    periodo = db.get(PeriodoAcademico, id_periodo)
    if not periodo or periodo.id_anio_lectivo != da.id_anio_lectivo:
        raise HTTPException(400, "Período inválido")

    docente = db.get(Usuario, da.id_usuario_docente)
    if not docente: raise HTTPException(404, "Docente no encontrado")

    estudiantes = db.execute(text("""
        SELECT p.apellido, p.nombre, c.calificacion_numerica
        FROM persona p
        JOIN matricula m ON m.id_persona = p.id_persona
        LEFT JOIN calificacion c ON c.id_persona = p.id_persona 
            AND c.id_asignatura = :asig AND c.id_periodo = :periodo 
            AND c.id_anio_lectivo = :anio
        WHERE m.id_grupo = :grupo AND m.id_anio_lectivo = :anio AND m.activo = TRUE
        ORDER BY p.apellido
    """), {
        "asig": da.id_asignatura, "periodo": id_periodo,
        "anio": da.id_anio_lectivo, "grupo": da.id_grupo
    }).fetchall()

    return {
        "asignatura": da.asignatura.nombre_asignatura,
        "docente": f"{docente.persona.nombre} {docente.persona.apellido}",
        "grupo": da.grupo.codigo_grupo,
        "periodo": periodo.nombre_periodo,
        "anio": da.anio_lectivo.anio,
        "estudiantes": [
            {
                "apellido": e.apellido,
                "nombre": e.nombre,
                "nota": round(e.calificacion_numerica, 1) if e.calificacion_numerica else None
            } for e in estudiantes
        ]
    }